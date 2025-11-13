#!/usr/bin/env python3
"""
Excel Formula Analyzer - Extracts and analyzes formulas from Excel Forecast sheet
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
import json
import re


def analyze_excel_formulas():
    """Analyze the Excel file and extract all formulas with their context"""
    
    file_path = "Corporate Modelling_230421.xlsx"
    print(f"Loading Excel file: {file_path}")
    
    # Load workbook with formulas preserved
    wb = load_workbook(file_path, data_only=False)
    ws = wb['Forecast']
    
    print(f"Forecast sheet dimensions: {ws.max_row} rows × {ws.max_column} columns\n")
    
    # Extract all formulas
    formulas = []
    data_values = {}
    
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            col_letter = openpyxl.utils.get_column_letter(col)
            cell_ref = f"{col_letter}{row}"
            
            # Get row label for context
            row_label = ws.cell(row=row, column=1).value or ""
            row_label = str(row_label).strip() if row_label else f"Row{row}"
            
            if cell.value is not None:
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    # It's a formula
                    formulas.append({
                        'cell': cell_ref,
                        'row': row,
                        'col': col,
                        'col_letter': col_letter,
                        'formula': cell.value,
                        'row_label': row_label
                    })
                else:
                    # It's a data value
                    data_values[cell_ref] = {
                        'value': cell.value,
                        'row_label': row_label
                    }
    
    # Categorize formulas
    categories = {
        'growth_rates': [],
        'ratios': [],
        'sums': [],
        'simple_arithmetic': [],
        'references': [],
        'other': []
    }
    
    for formula in formulas:
        f = formula['formula']
        
        if '^(' in f and '/1/' in f and '-1' in f:
            categories['growth_rates'].append(formula)
        elif 'SUM(' in f:
            categories['sums'].append(formula)
        elif '/' in f and not '^(' in f and not 'SUM(' in f:
            categories['ratios'].append(formula)
        elif any(op in f for op in ['+', '-', '*']) and not any(func in f for func in ['SUM(', '^(']):
            categories['simple_arithmetic'].append(formula)
        elif f.startswith('=$'):
            categories['references'].append(formula)
        else:
            categories['other'].append(formula)
    
    return {
        'formulas': formulas,
        'data_values': data_values,
        'categories': categories,
        'sheet_info': {
            'rows': ws.max_row,
            'columns': ws.max_column
        }
    }


if __name__ == "__main__":
    analysis = analyze_excel_formulas()
    
    print("=== FORMULA ANALYSIS SUMMARY ===")
    print(f"Total formulas found: {len(analysis['formulas'])}")
    print(f"Total data values: {len(analysis['data_values'])}")
    
    print("\nFormula categories:")
    for category, formulas in analysis['categories'].items():
        print(f"  {category}: {len(formulas)} formulas")
    
    print("\n=== SAMPLE FORMULAS BY CATEGORY ===")
    
    for category, formulas in analysis['categories'].items():
        if formulas:
            print(f"\n{category.upper()} (showing first 3):")
            for i, f in enumerate(formulas[:3]):
                print(f"  {f['cell']}: {f['formula']} | {f['row_label']}")
    
    # Save detailed analysis to JSON for further processing
    with open('formula_analysis.json', 'w') as f:
        json.dump(analysis, f, indent=2, default=str)
    
    print(f"\nDetailed analysis saved to 'formula_analysis.json'")